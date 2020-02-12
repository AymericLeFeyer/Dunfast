import os
import unicodedata
import numpy as np
import pyexcel as p

import unicodedata
from tkinter import *
from openpyxl import *
from openpyxl.styles import PatternFill

import interface

Lots = []


def start(NumSemaine, ws, C, self):
    tab = []
    tab.append(tryTo(DYNAMAN, NumSemaine, ws, C, self) - 1)
    tab.append(tryTo(FRET, NumSemaine, ws, C, self))
    tab.append(tryTo(CROSS_ST_DIRECT, NumSemaine, ws, C, self))
    tab.append(tryTo(CENTRE_EMPOTAGE, NumSemaine, ws, C, self))
    tab.append(tryTo(LOTS_BLOQUES, NumSemaine, ws, C, self))
    tab.append(tryTo(FENES, NumSemaine, ws, C, self))
    tab.append(tryTo(SCAFRUIT, NumSemaine, ws, C, self))
    tab.append(tryTo(COMMENTS, NumSemaine, ws, C, self))
    tab.append(tryTo(POLYBAG, NumSemaine, ws, C, self))
    tryTo(SCORES, NumSemaine, ws, C, self)

    return tab


# Appels des fonctions
def tryTo(f, NumSemaine, ws, C, self):
    try:
        a = f(NumSemaine, ws, C)
    except Exception as exception:
        interface.printException(self,
                                 "Il faut insérer le fichier dans le dossier " + str(f) + ". Erreur " + str(exception))
    else:
        print("Fichier " + str(f) + " accepté")
        return a


# Création des dossiers pour la semaine S
def createFolders(s, ws, C):
    print("Création des dossiers pour la semaine ...")
    semaine = "\Dunfast\Semaine " + str(s)
    user = os.path.expanduser('~')

    if not (os.path.isdir(user + "\Dunfast")):
        os.mkdir(user + "\Dunfast")

    if not (os.path.isdir(user + semaine)):
        os.mkdir(user + semaine)

    os.mkdir(user + semaine + "/FRET")
    os.mkdir(user + semaine + "/CROSS_ST_DIRECT")
    os.mkdir(user + semaine + "/FENES")
    os.mkdir(user + semaine + "/CENTRE_EMPOTAGE")
    os.mkdir(user + semaine + "/LOTS_BLOQUES")
    os.mkdir(user + semaine + "/SCAFRUIT")
    os.mkdir(user + semaine + "/COMMENTAIRES")
    os.mkdir(user + semaine + "/FICHIERS_DYNAMAN")
    os.mkdir(user + semaine + "/POLYBAG")
    os.mkdir(user + semaine + "/SCORES")
    os.mkdir(user + semaine + "/SMARTFRESH")
    CREATE_SCAFRUIT(s, ws, C, user + semaine)
    CREATE_COMMENTS(s, ws, C, user + semaine)
    print("Dossiers créés, vous pouvez y introduire les différents documents")


def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    only_ascii = nfkd_form.encode('ASCII', 'ignore')
    return only_ascii

def SCORES(NumSemaine, p1, C):
    a = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/SCORES")
    score1 = load_workbook(filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/SCORES/" + str(
            a[0]))

    score2 = load_workbook(filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/SCORES/" + str(
            a[1]))

    feuilleScore1 = score1.active
    feuilleScore2 = score2.active

    wb = Workbook()
    fn = "Scores.xlsx"
    ws = wb.active
    ws.title = "Scores"

    # Remplissage du tableau
    id = 1
    ws['A1'] = "idScore"
    ws['B1'] = "tonnage"
    ws['C1'] = "depotage"
    ws['D1'] = "container"
    ws['E1'] = "marque"
    ws['F1'] = "produit"
    ws['G1'] = "qteAnnonce"
    ws['H1'] = "numLot"
    ws['I1'] = "gestion"
    ws['J1'] = "prio"
    ws['K1'] = "info"
    ws['L1'] = "resultat"
    ws['M1'] = "mdc"
    ws['N1'] = "ncc"

    for index in range(4, feuilleScore1.max_row + 1):
         ws['A'+str(id+1)] = int(id)
         ws['B'+str(id+1)] = feuilleScore1['A'+str(index)].value
         ws['C'+str(id+1)] = feuilleScore1['B'+str(index)].value
         ws['D'+str(id+1)] = feuilleScore1['C'+str(index)].value
         ws['E'+str(id+1)] = feuilleScore1['D'+str(index)].value
         ws['F'+str(id+1)] = feuilleScore1['E'+str(index)].value
         ws['G'+str(id+1)] = feuilleScore1['F'+str(index)].value
         ws['H'+str(id+1)] = feuilleScore1['G'+str(index)].value
         ws['I'+str(id+1)] = feuilleScore1['H'+str(index)].value
         ws['J'+str(id+1)] = feuilleScore1['I'+str(index)].value
         ws['K'+str(id+1)] = feuilleScore1['J'+str(index)].value
         ws['L'+str(id+1)] = feuilleScore1['K'+str(index)].value
         ws['M'+str(id+1)] = feuilleScore1['L'+str(index)].value
         ws['N'+str(id+1)] = feuilleScore1['M'+str(index)].value
         id += 1
    for index in range(4, feuilleScore2.max_row + 1):
         ws['A'+str(id+1)] = int(id)
         ws['B'+str(id+1)] = feuilleScore2['A'+str(index)].value
         ws['C'+str(id+1)] = feuilleScore2['B'+str(index)].value
         ws['D'+str(id+1)] = feuilleScore2['C'+str(index)].value
         ws['E'+str(id+1)] = feuilleScore2['D'+str(index)].value
         ws['F'+str(id+1)] = feuilleScore2['E'+str(index)].value
         ws['G'+str(id+1)] = feuilleScore2['F'+str(index)].value
         ws['H'+str(id+1)] = feuilleScore2['G'+str(index)].value
         ws['I'+str(id+1)] = feuilleScore2['H'+str(index)].value
         ws['J'+str(id+1)] = feuilleScore2['I'+str(index)].value
         ws['K'+str(id+1)] = feuilleScore2['J'+str(index)].value
         ws['L'+str(id+1)] = feuilleScore2['K'+str(index)].value
         ws['M'+str(id+1)] = feuilleScore2['L'+str(index)].value
         ws['N'+str(id+1)] = feuilleScore2['M'+str(index)].value
         id += 1





    wb.save(filename=os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "\SMARTFRESH\Scores.xlsx")


    

def DYNAMAN(NumSemaine, ws, C):
    a = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FICHIERS_DYNAMAN")

    dyna1 = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FICHIERS_DYNAMAN/" + str(
            a[0]))
    dyna2 = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FICHIERS_DYNAMAN/" + str(
            a[1]))
    feuilleDyna1 = dyna1.active
    feuilleDyna2 = dyna2.active

    ws['D2'] = feuilleDyna2['E2'].value
    ws['E2'] = feuilleDyna2['A2'].value
    ws['C2'] = "SQ"

    j = 1
    fruidor = 0

    print(feuilleDyna2.max_row + 1)
    for index in range(2, feuilleDyna2.max_row + 1):
        if ws['D' + str(j)].value != feuilleDyna2['E' + str(index)].value:
            j += 1
            ws['D' + str(j)] = int(feuilleDyna2['F' + str(index)].value.replace("_x003", "").replace("_", ""))
            ws['E' + str(j)] = feuilleDyna2['E' + str(index)].value
            ws['C' + str(j)] = "SQ"
            Lots.append(ws['D' + str(j)].value)
            # C.fruidor.append(int(feuilleDyna1['F' + str(index)].value))
            fruidor += 1

    print(feuilleDyna1.max_row + 1)
    for index in range(2, feuilleDyna1.max_row + 1):
        if ws['D' + str(j)].value != feuilleDyna1['E' + str(index)].value:
            j += 1
            ws['D' + str(j)] = int(feuilleDyna1['F' + str(index)].value.replace("_x003", "").replace("_", ""))
            ws['E' + str(j)] = feuilleDyna1['E' + str(index)].value
            ws['C' + str(j)] = "SQ"
            Lots.append(ws['D' + str(j)].value)
            # C.fruidor.append(int(feuilleDyna1['F' + str(index)].value))
            fruidor += 1

    print("Il y a un total de " + str(fruidor) + " containers")

    return fruidor + 1


# Récupération des informations du fichier Fret
def FRET(NumSemaine, ws, C):
    a = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FRET")

    fret = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FRET/" + str(a[0]))
    feuilleFret = fret.active

    # Importations des informations du fichier Fret vers le fichier principal

    j = 2
    francit = 0

    for index in range(2, feuilleFret.max_row + 1):
        if remove_accents(str(feuilleFret['F' + str(index)].value)) == b'reserve francite':
            C.francite.append(int(feuilleFret['E' + str(index)].value))
            for j in range(2, ws.max_row + 1):
                if ws['D' + str(j)].value in C.francite:
                    if ws['B' + str(j)].value:
                        pass
                    else:
                        ws['B' + str(j)].value = "F"
                        francit += 1

    if a[0] != 'Fret.xlsx':
        os.rename(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FRET/" + a[0],
                  r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FRET/Fret.xlsx")

    return francit


# Récupération des informations dans le fichier CrossDock, ST, Direct
def CROSS_ST_DIRECT(NumSemaine, ws, C):
    cross_st_direct = 1
    b = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/CROSS_ST_DIRECT")
    cross = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/CROSS_ST_DIRECT/" + str(
            b[0]))
    feuilleCross = cross.active

    # Récupération des 3 catégories

    st = []
    direct = []
    cross = []

    for index in range(6, 100):
        if feuilleCross['A' + str(index)].value is not None:
            st.append(feuilleCross['A' + str(index)].value)
        if feuilleCross['C' + str(index)].value is not None:
            direct.append(feuilleCross['C' + str(index)].value)
        if feuilleCross['F' + str(index)].value is not None:
            cross.append(feuilleCross['F' + str(index)].value)

    # Importations des informations du fichier Cross vers le fichier principal

    for index in range(len(st)):
        if st[index] not in Lots:
            ws['D' + str(ws.max_row + 1)] = st[index]
            ws['F' + str(ws.max_row)].value = ""
            Lots.append(st[index])
    for index in range(len(direct)):
        if direct[index] not in Lots:
            ws['D' + str(ws.max_row + 1)] = direct[index]
            ws['F' + str(ws.max_row)].value = ""
            Lots.append(direct[index])
    for index in range(len(cross)):
        if cross[index] not in Lots:
            ws['D' + str(ws.max_row + 1)] = cross[index]
            ws['F' + str(ws.max_row)].value = ""
            Lots.append(cross[index])

    for index in range(2, ws.max_row + 1):
        if st:
            if ws['D' + str(index)].value in st:
                if ws['F' + str(index)].value:
                    ws['F' + str(index)].value += "Sans tri"
                else:
                    ws['F' + str(index)].value = "Sans tri"
                ws['C' + str(index)].value = "SQ"
        if direct:
            if ws['D' + str(index)].value in direct:
                if ws['F' + str(index)].value:
                    ws['F' + str(index)].value += "Direct"
                else:
                    ws['F' + str(index)].value = "Direct"
                ws['C' + str(index)].value = "SQ"
        if cross:
            if ws['D' + str(index)].value in cross:
                if ws['F' + str(index)].value:
                    ws['F' + str(index)].value += "CrossDock"
                else:
                    ws['F' + str(index)].value = "CrossDock"
                ws['C' + str(index)].value = "SQ"

    return [len(st), len(cross), len(direct)]


# Récupération des informations pour le centre d'empotage
def CENTRE_EMPOTAGE(NumSemaine, ws, C):
    c = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/CENTRE_EMPOTAGE")
    zoneC = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/CENTRE_EMPOTAGE/" + str(
            c[0]))
    feuilleZoneC = zoneC.active

    totalC = 0

    # Importations des informations du fichier Centre d'Empotage vers le fichier principal

    dataZoneC = [[cell.value for cell in row] for row in feuilleZoneC.rows]
    betterDataZoneC = np.ravel(dataZoneC)
    for i in range(len(betterDataZoneC)):
        if betterDataZoneC[i]:
            betterDataZoneC[i] = int(str(betterDataZoneC[i]).replace("_", ""))

    for index in range(2, ws.max_row):
        if ws['D' + str(index)].value in betterDataZoneC:
            if ws['A' + str(index)].value:
                pass
            else:
                ws['A' + str(index)].value = "C"
                totalC += 1

    print("Il y a " + str(totalC) + " zone c")

    return totalC


# Récupération des informations pour les lots bloqués
def LOTS_BLOQUES(NumSemaine, ws, C):
    d = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/LOTS_BLOQUES")
    bloquer = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/LOTS_BLOQUES/" + str(d[0]))
    feuilleBloquer = bloquer.active

    greyFill = PatternFill(start_color='969696',
                           end_color='969696',
                           fill_type='solid')

    # Importations des informations du fichier Lots bloqués vers le fichier principal

    dataBloquer = []
    for i in range(2, feuilleBloquer.max_row + 1):
        dataBloquer.append(feuilleBloquer['A' + str(i)].value)

    for i in range(2, ws.max_row + 1):
        if ws['D' + str(i)].value in dataBloquer:
            ws['G' + str(i)].value = "Oui"

    return True


# Récupérations des informations pour les contremarques spécifiques (fenes)
def FENES(NumSemaine, ws, C):
    e = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FENES")
    spe = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/FENES/" + str(e[0]))
    feuilleFenes = spe.active

    # Importations des informations du fichier Fenes vers le fichier principal

    lotsFenes = []
    letter = 'B'
    for i in range(2, feuilleFenes.max_column):
        lotsFenes.append(feuilleFenes[letter + str(4)].value)
        lotsFenes.append(feuilleFenes[letter + str(19)].value)
        letter = chr(ord(letter) + 1)

    for i in range(2, ws.max_row):
        if ws['D' + str(i)].value in lotsFenes:
            if ws['F' + str(i)].value is None:
                ws['F' + str(i)].value = "Contremarque spé "
            else:
                ws['F' + str(i)].value += " + Contremarque spé "

    return True


# Création du fichier Scafruit, à modifier par Dunfresh
def CREATE_SCAFRUIT(NumSemaine, ws, C, path):
    sf = Workbook()
    ss = sf.active
    ss.title = "Scafruit"
    ss.column_dimensions['A'].width = 30
    ss.column_dimensions['B'].width = 30
    ss.column_dimensions['C'].width = 30
    ss.column_dimensions['D'].width = 30
    ss.column_dimensions['F'].width = 30
    ss.column_dimensions['G'].width = 30
    ss.column_dimensions['H'].width = 30

    ss['A1'].value = "Numéro du lot"
    ss['B1'].value = "Quantité"
    ss['C1'].value = "Catégorie"
    ss['D1'].value = "Par combien ?"
    ss['E1'].value = "+"
    ss['F1'].value = "Quantité"
    ss['G1'].value = "Catégorie ?"
    ss['H1'].value = "Par combien ?"

    sf.save(path + '/SCAFRUIT/Scafruit.xlsx')


# Récupération des informations pour le Scafruit
def SCAFRUIT(NumSemaine, ws, C):
    f = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/SCAFRUIT")
    scaf = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/SCAFRUIT/" + str(f[0]))
    feuilleScafruit = scaf.active
    Lot = []
    Qte = []
    Cat = []
    How = []
    ind = 0

    # Récupération des informations du fichier Scafruit.xlsx

    for i in range(2, feuilleScafruit.max_row + 1):
        Lot.append(int(str(feuilleScafruit['A' + str(i)].value).replace("_", "")))
        Qte.append(feuilleScafruit['B' + str(i)].value)
        Cat.append(feuilleScafruit['C' + str(i)].value)
        How.append(str(feuilleScafruit['D' + str(i)].value))

        if feuilleScafruit['F' + str(i)].value:
            How[len(How) - 1] += " + " + str(feuilleScafruit['F' + str(i)].value) + " " + str(
                feuilleScafruit['G' + str(i)].value) + " en " + str(feuilleScafruit['H' + str(i)].value)

    for j in range(2, ws.max_row + 1):
        if ws['D' + str(j)].value in Lot:

            ind = Lot.index(ws['D' + str(j)].value)
            C.scafruit.append(Lot[ind])

            if ws['F' + str(j)].value is None:
                ws['F' + str(j)].value = "SCAFRUIT " + str(Qte[ind]) + " " + str(Cat[ind]) + " en " + str(How[ind])
            else:
                ws['F' + str(j)].value += "+ SCAFRUIT " + str(Qte[ind]) + " " + str(Cat[ind]) + " en " + str(How[ind])

    return len(Lot)


# Création du fichier Commentaires, à modifier par Dunfresh
def CREATE_COMMENTS(NumSemaine, ws, C, path):
    pr = Workbook()
    ps = pr.active
    ps.title = "Commentaires"
    ps.column_dimensions['A'].width = 30
    ps.column_dimensions['B'].width = 50
    ps.column_dimensions['C'].width = 15
    ps.column_dimensions['D'].width = 15
    ps.column_dimensions['E'].width = 15
    ps.column_dimensions['F'].width = 15
    ps.column_dimensions['G'].width = 15

    ps['A1'].value = "Numéro du lot"
    ps['B1'].value = "Commentaire"
    ps['C1'].value = "Appel SQ"
    ps['D1'].value = "Nexy"
    # ps['E1'].value = "Polybag orange"
    # ps['F1'].value = "Polybag complet"
    ps['G1'].value = "Prio"

    for i in range(2, 200):
        ps['A' + str(i)].value = int(i - 1)

    pr.save(path + '/COMMENTAIRES/Commentaires.xlsx')


# Récupération des informations pour les Commentaires
def COMMENTS(NumSemaine, ws, C):
    g = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/COMMENTAIRES")
    com = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/COMMENTAIRES/" + str(g[0]))
    feuilleComments = com.active

    tableau = [0, 0, 0, 0]

    for i in range(2, feuilleComments.max_row + 1):
        a = feuilleComments['A' + str(i)].value

        b = feuilleComments['B' + str(i)].value

        if b:
            addComment(a, b, ws)

        b = feuilleComments['C' + str(i)].value

        if b:
            addComment(a, "Appel SQ ", ws)
            tableau[0] += 1

        b = feuilleComments['D' + str(i)].value

        if b:
            addComment(a, "Nexy ", ws)
            tableau[1] += 1

        # b = feuilleComments['E' + str(i)].value
        # if b:
        #     addComment(a, "Poly orange ", ws)
        #     tableau[2] += 1
        #
        # b = feuilleComments['F' + str(i)].value
        # if b:
        #     addComment(a, "Poly complet ", ws)
        #     tableau[2] += 1

        b = feuilleComments['G' + str(i)].value
        if b:
            addComment(a, "Prio ", ws)
            tableau[3] += 1

    return tableau


# Ajouter les commentaires associes
def addComment(n, c, ws):
    w = 0
    if n in Lots:
        for i in range(2, ws.max_row + 1):
            if int(ws['D' + str(i)].value) == int(n):
                w = i
                break

        if ws['F' + str(w)].value:
            ws['F' + str(w)].value += " + " + str(c)
        else:
            ws['F' + str(w)].value = str(c)


# Polybags
def POLYBAG(NumSemaine, ws, C):
    p = os.listdir(r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/POLYBAG")
    poly = load_workbook(
        filename=r"" + os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "/POLYBAG/" + str(p[0]))
    feuillePoly = poly.active
    nb = 0

    containersPoly = []
    letter = 'A'
    for j in range(feuillePoly.max_column):

        for i in range(2, feuillePoly.max_row + 1):
            if feuillePoly[chr(ord(letter) + 1) + str(i)].value:
                containersPoly.append(
                    [feuillePoly[letter + str(i)].value, feuillePoly[chr(ord(letter) + 1) + str(i)].value])
        letter = chr(int(ord(letter) + 2))
        j += 1

    for i in range(len(containersPoly)):
        addComment(int(containersPoly[i][0]), "Polybag " + str(containersPoly[i][1]), ws)
        nb += 1

    return nb


def saveFile(num, wb2):
    wb2.save(filename=os.path.expanduser('~') + "\Dunfast\Semaine " + str(num) + "/Antilles " + str(num) + ".xlsx")
    if num == 0:
        pass
