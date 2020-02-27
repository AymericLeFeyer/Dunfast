import openpyxl
import os
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from openpyxl import *

Tab = []


class AllContainer:
    def __init__(self):
        self.centre_empotage = ''
        self.francite = ''
        self.sq = ''
        self.lot = 0
        self.num = ''
        self.bloque = False
        self.commentaire = ''


def createNewTab(ws, ws2, NumSemaine):
    takeInfos(ws)
    createTab(ws2)
    createContainerTab(NumSemaine)


def takeInfos(ws):
    for i in range(2, ws.max_row + 1):
        a = AllContainer()

        # Centre d'empotage
        if ws['A' + str(i)].value:
            a.centre_empotage = 'C'

        # Francite
        if ws['B' + str(i)].value:
            a.francite = 'F'

        # SQ
        if ws['C' + str(i)].value:
            a.sq = 'SQ'

        # Numero Lot
        a.lot = int(ws['D' + str(i)].value)

        # Numero container
        a.num = str(ws['E' + str(i)].value)

        # A bloquer
        if ws['G' + str(i)].value:
            a.bloque = True

        # Commentaires
        if ws['F' + str(i)].value:
            a.commentaire = ws['F' + str(i)].value

        Tab.append(a)


def putSomeStyle(ws2, numColonne, greyFill):
    font = Font(name='Calibri',
                size=17,
                bold=True,
                italic=False,
                vertAlign='baseline',
                underline='none',
                strike=False,
                color='FF000000')

    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))

    centre = Alignment(horizontal='centerContinuous')

    ws2.merge_cells(str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1) + ":" + str(
        openpyxl.utils.cell.get_column_letter(numColonne + 5)) + str(1)))
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(1)].font = font
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(1)].fill = greyFill
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(1)].border = medium_border
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(1)].alignment = centre
    ws2.row_dimensions[1].height = 20
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne))].width = 3
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne + 1))].width = 3
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne + 2))].width = 4
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne + 3))].width = 6
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne + 4))].width = 15
    ws2.column_dimensions[str(openpyxl.utils.cell.get_column_letter(numColonne + 5))].width = 60


def fillTableau(ws2, numColonne, a, Tablo, greyFill, i):
    thin_boder = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(a)].value = Tablo[i - 2].francite
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne)) + str(a)].border = thin_boder
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 1)) + str(a)].value = Tablo[
        i - 2].centre_empotage
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 1)) + str(a)].border = thin_boder
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 2)) + str(a)].value = Tablo[i - 2].sq
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 2)) + str(a)].border = thin_boder
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 3)) + str(a)].value = Tablo[i - 2].lot
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 3)) + str(a)].border = thin_boder
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 4)) + str(a)].value = Tablo[i - 2].num
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 4)) + str(a)].border = thin_boder
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 5)) + str(a)].value = Tablo[i - 2].commentaire
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 5)) + str(a)].border = thin_boder
    if Tablo[i - 2].bloque:
        ws2[str(openpyxl.utils.cell.get_column_letter(numColonne + 4)) + str(a)].fill = greyFill

def createContainerTab(NumSemaine):
    Tableau = sorted(Tab, key=lambda container: container.num)
    wb = Workbook()
    fn = "Containers.xlsx"
    ws = wb.active
    ws.title = "Containers"
    ws['A1'] = "id"
    ws['B1'] = "numContainer"
    ws['C1'] = "isSQ"
    ws['D1'] = "isC"
    ws['E1'] = "isF"
    ws['F1'] = "isBloque"
    ws['G1'] = "commentaires"

    for i in range(len(Tableau)):
        ws['A'+str(i+2)] = Tableau[i].lot
        ws['B'+str(i+2)] = Tableau[i].num
        if Tableau[i].sq:
            ws['C'+str(i+2)] = 1
        if Tableau[i].centre_empotage:
            ws['D'+str(i+2)] = 1
        if Tableau[i].francite:
            ws['E'+str(i+2)] = 1
        if Tableau[i].bloque:
            ws['F'+str(i+2)] = 1
        ws['G'+str(i+2)] = Tableau[i].commentaire
        





    wb.save(filename=os.path.expanduser('~') + "\Dunfast\Semaine " + str(NumSemaine) + "\SMARTFRESH\Containers.xlsx")


def createTab(ws2):
    greyFill = PatternFill(start_color='969696',
                           end_color='969696',
                           fill_type='solid')

    Tableau = sorted(Tab, key=lambda container: container.num)
    TabFrancite = []
    TabZoneA = []
    TabScafruit = []
    for c in Tableau:
        if 'SCAFRUIT' in c.commentaire:
            TabScafruit.append(c)
        elif c.francite == 'F':
            TabFrancite.append(c)
        else:
            TabZoneA.append(c)

    # On fait des colonnes de 60 lignes max
    print(len(TabFrancite), len(TabScafruit), len(TabZoneA))

    numColonne = 1
    a = 2
    putSomeStyle(ws2, numColonne, greyFill)
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Francité"

    for i in range(2, len(TabFrancite) + 2):
        if a % 60 == 0:
            numColonne += 6
            a = 2
            putSomeStyle(ws2, numColonne, greyFill)
            ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Francité"
        fillTableau(ws2, numColonne, a, TabFrancite, greyFill, i)

        a += 1

    numColonne += 6
    a = 2
    putSomeStyle(ws2, numColonne, greyFill)
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Zone A"

    for i in range(2, len(TabZoneA) + 2):
        if a % 60 == 0:
            numColonne += 6
            a = 2
            putSomeStyle(ws2, numColonne, greyFill)
            ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Zone A"

        fillTableau(ws2, numColonne, a, TabZoneA, greyFill, i)

        a += 1

    numColonne += 6
    a = 2
    putSomeStyle(ws2, numColonne, greyFill)
    ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Palettes à descendre"

    for i in range(2, len(TabScafruit) + 2):
        if a % 60 == 0:
            numColonne += 6
            a = 2
            putSomeStyle(ws2, numColonne, greyFill)
            ws2[str(openpyxl.utils.cell.get_column_letter(numColonne) + str(1))].value = "Palettes à descendre"

        fillTableau(ws2, numColonne, a, TabScafruit, greyFill, i)

        a += 1
